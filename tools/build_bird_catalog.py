"""Build the bundled global-bird catalog from upstream authoritative sources.

This is a *one-time* data-preparation script. Its outputs are committed into the
repository under ``analyzer/models/birds/`` and bundled with the app. We do NOT
fetch or rebuild this data at build time -- the runtime only reads the CSV.

Inputs (downloaded out-of-band, paths supplied via CLI flags):

  * IOC World Bird List (v15.1+) master XLSX
        Frank Gill, David Donsker & Pamela Rasmussen (Eds), licensed CC-BY 3.0
        https://www.worldbirdnames.org/

  * IBP-AOS Alpha Codes (Pyle & DeSante / Institute for Bird Populations) PDF
        4- and 6-letter standardized banding codes for North American birds
        https://www.birdpop.org/

Inputs read from the repository (committed):

  * analyzer/models/labels.txt                -- 500 model species (immutable)
  * analyzer/models/labels_scispecies.csv     -- model species -> scientific family
  * analyzer/models/scispecies_dispname.csv   -- scientific family -> display name

Outputs (committed):

  * analyzer/models/birds/birds_global.csv    -- master catalog
  * analyzer/models/birds/NOTICES.md          -- attribution for bundled data

Run:
  python tools/build_bird_catalog.py \\
      --ioc /tmp/ioc_v15.1.xlsx \\
      --alpha-codes-text /tmp/alpha_codes.txt
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
MODELS_DIR = REPO_ROOT / 'analyzer' / 'models'
OUT_DIR = MODELS_DIR / 'birds'


# Hand-curated overrides for the 34 AOS model labels whose preferred English
# name differs from IOC v15.1. Verified one-by-one against IOC scientific
# binomials (see ``tools/build_bird_catalog.py`` history). Each entry maps:
#   model AOS common name -> (IOC scientific binomial, IOC English name)
# The IOC scientific binomial is the authoritative join key; the IOC English
# name is preserved as a search alias.
AOS_TO_IOC_OVERRIDES: dict[str, tuple[str, str]] = {
    'Bank Swallow':                    ('Riparia riparia',          'Sand Martin'),
    'Barn Owl':                        ('Tyto alba',                'Western Barn Owl'),
    'Black Swift':                     ('Cypseloides niger',        'American Black Swift'),
    'Black-bellied Plover':            ('Pluvialis squatarola',     'Grey Plover'),
    'Black-bellied Whistling-Duck':    ('Dendrocygna autumnalis',   'Black-bellied Whistling Duck'),
    'Black-crowned Night-Heron':       ('Nycticorax nycticorax',    'Black-crowned Night Heron'),
    'Brant':                           ('Branta bernicla',          'Brant Goose'),
    'Bushtit':                         ('Psaltriparus minimus',     'American Bushtit'),
    'Cattle Egret':                    ('Ardea ibis',               'Western Cattle Egret'),
    'Chukar':                          ('Alectoris chukar',         'Chukar Partridge'),
    'Cliff Swallow':                   ('Petrochelidon pyrrhonota', 'American Cliff Swallow'),
    'Common Raven':                    ('Corvus corax',             'Northern Raven'),
    'Common Redpoll':                  ('Acanthis flammea',         'Redpoll'),
    'Dovekie':                         ('Alle alle',                'Little Auk'),
    'Dusky Flycatcher':                ('Empidonax oberholseri',    'American Dusky Flycatcher'),
    'Eared Grebe':                     ('Podiceps nigricollis',     'Black-necked Grebe'),
    'European Starling':               ('Sturnus vulgaris',         'Common Starling'),
    'Fox Sparrow':                     ('Passerella iliaca',        'Red Fox Sparrow'),
    'Gray-crowned Rosy-Finch':         ('Leucosticte tephrocotis',  'Grey-crowned Rosy Finch'),
    'Herring Gull':                    ('Larus smithsonianus',      'American Herring Gull'),
    # AOS still recognizes Hoary Redpoll; IOC v15.1 lumps with Common Redpoll
    # (Acanthis flammea). We retain the AOS scientific name so the two model
    # labels stay distinguishable in the catalog.
    'Hoary Redpoll':                   ('Acanthis hornemanni',      'Arctic Redpoll'),
    'House Wren':                      ('Troglodytes aedon',        'Northern House Wren'),
    'Northern Goshawk':                ('Astur atricapillus',       'American Goshawk'),
    'Northern Hawk Owl':               ('Surnia ulula',             'Northern Hawk-Owl'),
    'Pacific-slope Flycatcher':        ('Empidonax difficilis',     'Western Flycatcher'),
    'Ring-necked Pheasant':            ('Phasianus colchicus',      'Common Pheasant'),
    'Rock Pigeon':                     ('Columba livia',            'Rock Dove'),
    'Rough-legged Hawk':               ('Buteo lagopus',            'Rough-legged Buzzard'),
    'Whimbrel':                        ('Numenius hudsonicus',      'Hudsonian Whimbrel'),
    'White Ibis':                      ('Eudocimus albus',          'American White Ibis'),
    'White-winged Crossbill':          ('Loxia leucoptera',         'Two-barred Crossbill'),
    'Yellow Warbler':                  ('Setophaga aestiva',        'American Yellow Warbler'),
    'Yellow-crowned Night-Heron':      ('Nyctanassa violacea',      'Yellow-crowned Night Heron'),
    'Yellow-rumped Warbler':           ('Setophaga coronata',       'Myrtle Warbler'),
}

# Light spelling-only normalisations -- applied when the override table doesn't
# match and the only difference is British vs American spelling or hyphenation.
# Order matters: we try the variants in sequence.
def _spelling_variants(name: str) -> list[str]:
    out = [name]
    if 'Gray' in name:
        out.append(name.replace('Gray', 'Grey'))
    if 'gray' in name:
        out.append(name.replace('gray', 'grey'))
    if '-' in name:
        out.append(name.replace('-', ' '))
    if 'Gray' in name and '-' in name:
        out.append(name.replace('Gray', 'Grey').replace('-', ' '))
    return out


REGION_CODES = {
    'NA':        'North America',
    'MA':        'Middle America',
    'SA':        'South America',
    'AF':        'Africa',
    'PAL':       'Eurasia (Palearctic)',
    'OR':        'Oriental (S/SE Asia)',
    'AU':        'Australasia',
    'AN':        'Antarctic',
    'AO':        'Atlantic Ocean',
    'PO':        'Pacific Ocean',
    'IO':        'Indian Ocean',
    'SO':        'Southern Ocean',
    'TrO':       'Tropical Ocean',
    'Worldwide': 'Worldwide',
}


def parse_ioc(xlsx_path: Path) -> list[dict]:
    """Parse the IOC master XLSX. Returns a list of species-level dicts.

    Forward-fills Order/Family/Genus across the sparse worksheet layout.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl required to parse IOC XLSX: pip install openpyxl")

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb['Master']

    curr = {'order': None, 'family_sci': None, 'family_eng': None, 'genus': None}
    species: list[dict] = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        # Columns (0-indexed): 2 Order, 3 Family(Sci), 4 Family(Eng), 5 Genus,
        # 6 Species(Sci epithet), 7 Subspecies, 9 Species(Eng), 10 Breeding Range
        if r[2]: curr['order'] = str(r[2]).title()
        if r[3]: curr['family_sci'] = str(r[3])
        if r[4]: curr['family_eng'] = str(r[4])
        if r[5]: curr['genus'] = str(r[5])
        if r[6] and not r[7]:
            sci = f"{curr['genus']} {r[6]}"
            species.append({
                'common_name': (r[9] or '').strip(),
                'scientific_name': sci,
                'family_sci': curr['family_sci'] or '',
                'family_common': curr['family_eng'] or '',
                'order': curr['order'] or '',
                'breeding_range': str(r[10] or '').strip(),
            })
    return species


_ALPHA_PAT = re.compile(
    r'^\s*(?:(\+)\s+)?'                        # optional '+' (non-species marker)
    r'([A-Za-z][A-Za-z0-9\'\-/ \.]+?)'         # english name (lazy)
    r'\s{2,}([A-Z]{4}\*?)'                     # 4-letter code (optional *)
    r'\s{2,}([A-Za-z][A-Za-z0-9\'\-/x\. ]+?)'  # scientific name
    r'\s{2,}([A-Z]{6}\*?)\s*$'                 # 6-letter code (optional *)
)


def parse_alpha_codes_text(txt_path: Path) -> list[dict]:
    """Parse pdftotext -layout output of the IBP alpha codes PDF."""
    rows: list[dict] = []
    with open(txt_path, encoding='utf-8') as f:
        for line in f:
            m = _ALPHA_PAT.match(line.rstrip('\n'))
            if not m:
                continue
            plus, eng, code4, sci, code6 = m.groups()
            if plus:
                # Skip non-species (morphs, hybrids, "Unidentified X")
                continue
            rows.append({
                'english': eng.strip(),
                'code4': code4.rstrip('*'),
                'scientific': sci.strip(),
            })
    return rows


def _read_model_labels(models_dir: Path) -> list[str]:
    with open(models_dir / 'labels.txt', encoding='utf-8-sig') as f:
        return [ln.strip() for ln in f if ln.strip()]


def _read_scispecies(models_dir: Path) -> dict[str, str]:
    """Returns ``{ model_species_name: scientific_family }``."""
    out: dict[str, str] = {}
    with open(models_dir / 'labels_scispecies.csv', encoding='utf-8-sig', newline='') as f:
        for row in csv.DictReader(f):
            sp = (row.get('Species') or '').strip()
            fam = (row.get('Scientific Family') or '').strip()
            if sp and fam:
                out[sp] = fam
    return out


def _read_family_display(models_dir: Path) -> dict[str, str]:
    """Returns ``{ scientific_family: display_name }`` from the existing CSV."""
    out: dict[str, str] = {}
    with open(models_dir / 'scispecies_dispname.csv', encoding='utf-8-sig', newline='') as f:
        for row in csv.DictReader(f):
            sci = (row.get('Scientific Family') or '').strip()
            disp = (row.get('Display Name') or '').strip()
            if sci and disp:
                out[sci] = disp
    return out


def _normalize_range(s: str) -> list[str]:
    """Split IOC ``Breeding Range`` cell into clean region codes."""
    if not s:
        return []
    parts = re.split(r'[,/&;]', s)
    out: list[str] = []
    for p in parts:
        tok = p.strip()
        if tok in REGION_CODES:
            out.append(tok)
        elif tok.lower().startswith('worldwide'):
            out.append('Worldwide')
    # Deduplicate while preserving order
    seen: set[str] = set()
    deduped: list[str] = []
    for tok in out:
        if tok not in seen:
            seen.add(tok)
            deduped.append(tok)
    return deduped


def build(ioc_xlsx: Path, alpha_codes_txt: Path) -> None:
    print(f"Reading IOC list from {ioc_xlsx}")
    ioc = parse_ioc(ioc_xlsx)
    print(f"  -> {len(ioc)} species")

    print(f"Reading alpha codes from {alpha_codes_txt}")
    alpha = parse_alpha_codes_text(alpha_codes_txt)
    print(f"  -> {len(alpha)} species with 4-letter codes")
    alpha_by_sci = {r['scientific']: r['code4'] for r in alpha}
    alpha_by_eng = {r['english']: r['code4'] for r in alpha}

    print("Reading existing model data")
    model_labels = _read_model_labels(MODELS_DIR)
    model_to_family_sci = _read_scispecies(MODELS_DIR)
    family_display = _read_family_display(MODELS_DIR)
    print(f"  -> {len(model_labels)} model labels, {len(family_display)} family display names")

    ioc_by_common = {s['common_name']: s for s in ioc}
    ioc_by_sci    = {s['scientific_name']: s for s in ioc}

    # Resolve each model label to an IOC species row.
    resolved: dict[str, dict] = {}   # canonical model name -> IOC row + aliases
    unresolved: list[str] = []
    fallback_synth: list[dict] = []  # for model species not in IOC at all (Hoary Redpoll)

    for label in model_labels:
        ioc_row: dict | None = None
        ioc_english_alias: str | None = None

        # 1. Direct match
        if label in ioc_by_common:
            ioc_row = ioc_by_common[label]
        # 2. Spelling variants (Gray<->Grey, hyphens)
        if ioc_row is None:
            for v in _spelling_variants(label):
                if v in ioc_by_common:
                    ioc_row = ioc_by_common[v]
                    ioc_english_alias = v if v != label else None
                    break
        # 3. Hand-curated override
        if ioc_row is None and label in AOS_TO_IOC_OVERRIDES:
            sci, eng = AOS_TO_IOC_OVERRIDES[label]
            ioc_row = ioc_by_sci.get(sci)
            ioc_english_alias = eng
            if ioc_row is None:
                # Override points to a scientific name IOC doesn't recognize
                # (e.g. Acanthis hornemanni -- IOC lumps it). Synthesize a
                # catalog row using the AOS data + existing family mapping.
                fallback_synth.append({
                    'label': label,
                    'sci': sci,
                    'ioc_eng': eng,
                })
                continue
        if ioc_row is None:
            unresolved.append(label)
            continue
        resolved[label] = {'ioc': ioc_row, 'alias': ioc_english_alias}

    if unresolved:
        print(f"ERROR: {len(unresolved)} model labels could not be mapped to IOC:")
        for u in unresolved:
            print(f"  - {u}")
        sys.exit(2)

    # Sanity: print which model labels were resolved via aliases / synth.
    via_alias = [k for k, v in resolved.items() if v['alias']]
    print(f"  model labels mapped directly:    {len(resolved) - len(via_alias)}")
    print(f"  model labels mapped via alias:   {len(via_alias)}")
    print(f"  model labels via synthetic row:  {len(fallback_synth)}")

    # Build the catalog. Use scientific name as dedup key. For model species we
    # prefer the AOS canonical (so existing tags don't move), and append the
    # IOC English name as a search alias.
    by_sci: dict[str, dict] = {}

    # First pass: insert all IOC species with IOC English as canonical.
    for row in ioc:
        sci = row['scientific_name']
        if sci in by_sci:
            continue
        by_sci[sci] = {
            'canonical': row['common_name'],
            'scientific': sci,
            'family_sci': row['family_sci'],
            'family_common_ioc': row['family_common'],
            'order': row['order'],
            'regions': ','.join(_normalize_range(row['breeding_range'])),
            'alpha_4': alpha_by_sci.get(sci, '') or alpha_by_eng.get(row['common_name'], ''),
            'aliases': '',
            'is_model_species': False,
        }

    # Second pass: override canonical name + flag is_model_species for matches.
    for model_label, info in resolved.items():
        sci = info['ioc']['scientific_name']
        entry = by_sci[sci]
        # Preserve the existing AOS-style canonical name from the model
        prior_canonical = entry['canonical']
        entry['canonical'] = model_label
        # Add IOC English (or the IOC-spelling variant) as search alias
        aliases = []
        if info['alias'] and info['alias'] != model_label:
            aliases.append(info['alias'])
        if prior_canonical and prior_canonical != model_label and prior_canonical not in aliases:
            aliases.append(prior_canonical)
        if aliases:
            entry['aliases'] = '|'.join(aliases)
        entry['is_model_species'] = True
        # Refresh alpha code lookup using model name too (IBP keys by AOS name)
        if not entry['alpha_4']:
            entry['alpha_4'] = alpha_by_eng.get(model_label, '')

    # Third pass: insert synthetic rows for model species IOC doesn't have.
    for synth in fallback_synth:
        fam_sci = model_to_family_sci.get(synth['label'], '')
        # IOC English family name + order -- find any IOC entry in the same family
        ioc_fam_eng = ''
        ioc_order = ''
        for row in ioc:
            if row['family_sci'].lower() == fam_sci.lower():
                ioc_fam_eng = row['family_common']
                ioc_order = row['order']
                break
        by_sci[synth['sci']] = {
            'canonical': synth['label'],
            'scientific': synth['sci'],
            'family_sci': fam_sci,
            'family_common_ioc': ioc_fam_eng,
            'order': ioc_order,
            'regions': 'NA',  # AOS-only labels are NA-centric
            'alpha_4': alpha_by_eng.get(synth['label'], ''),
            'aliases': synth['ioc_eng'] if synth['ioc_eng'] != synth['label'] else '',
            'is_model_species': True,
        }

    # Resolve family_common: prefer existing scispecies_dispname.csv mapping
    # (preserves "Thrush sp."-style display names), then IOC English family.
    for entry in by_sci.values():
        fam_sci = entry['family_sci']
        disp = family_display.get(fam_sci)
        if disp:
            entry['family_common'] = disp
        else:
            entry['family_common'] = entry['family_common_ioc']
        entry.pop('family_common_ioc', None)

    # Sanity: every model label is in the catalog.
    by_canonical = {e['canonical']: e for e in by_sci.values()}
    missing = [lbl for lbl in model_labels if lbl not in by_canonical]
    if missing:
        print(f"ERROR: {len(missing)} model labels missing from catalog after build:")
        for m in missing[:20]:
            print(f"  - {m}")
        sys.exit(3)

    # Emit CSV. Sort by canonical name for stable diffs.
    entries = sorted(by_sci.values(), key=lambda e: e['canonical'].lower())
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_csv = OUT_DIR / 'birds_global.csv'
    with open(out_csv, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=[
            'canonical_common_name', 'scientific_name', 'family_sci', 'family_common',
            'order', 'regions', 'alpha_4', 'aliases', 'is_model_species',
        ])
        w.writeheader()
        for e in entries:
            w.writerow({
                'canonical_common_name': e['canonical'],
                'scientific_name': e['scientific'],
                'family_sci': e['family_sci'],
                'family_common': e['family_common'],
                'order': e['order'],
                'regions': e['regions'],
                'alpha_4': e['alpha_4'],
                'aliases': e['aliases'],
                'is_model_species': '1' if e['is_model_species'] else '0',
            })
    print(f"Wrote {len(entries)} catalog rows -> {out_csv}")

    n_model = sum(1 for e in entries if e['is_model_species'])
    n_alpha = sum(1 for e in entries if e['alpha_4'])
    print(f"  rows flagged is_model_species: {n_model}")
    print(f"  rows with 4-letter code:       {n_alpha}")

    # NOTICES.md attribution.
    notices = OUT_DIR / 'NOTICES.md'
    notices.write_text(_render_notices(), encoding='utf-8')
    print(f"Wrote attribution -> {notices}")


def _render_notices() -> str:
    return (
        "# Bundled bird-catalog data sources\n\n"
        "The catalog file ``birds_global.csv`` is built from the following\n"
        "authoritative sources. Attribution required by license / convention is\n"
        "listed below.\n\n"
        "## IOC World Bird List (v15.1)\n\n"
        "Frank Gill, David Donsker & Pamela Rasmussen (Eds). 2025. "
        "*IOC World Bird List* (v15.1). doi:10.14344/IOC.ML.15.1. "
        "https://www.worldbirdnames.org/\n\n"
        "Licensed under "
        "[Creative Commons Attribution 3.0 Unported](https://creativecommons.org/licenses/by/3.0/).\n"
        "Common names, scientific binomials, taxonomic order/family/genus, and "
        "breeding-range biogeographic codes are derived from this source.\n\n"
        "## IBP-AOS Alpha Codes\n\n"
        "Pyle, P. and DeSante, D.F. *Four-letter (English Name) and Six-letter "
        "(Scientific Name) Alpha Codes for North American Birds.* "
        "The Institute for Bird Populations. https://www.birdpop.org/\n\n"
        "Per the 66th AOS Supplement (2025). 4-letter codes are reproduced as\n"
        "factual abbreviations; full attribution is preserved here in lieu of a\n"
        "publicly documented license.\n\n"
        "## ProjectKestrel additions\n\n"
        "* Hand-curated AOS-to-IOC name overrides for model species whose\n"
        "  preferred English name differs between authorities (see\n"
        "  ``tools/build_bird_catalog.py``, ``AOS_TO_IOC_OVERRIDES``).\n"
        "* Family display names (``family_common``) preserve the existing "
        "``analyzer/models/scispecies_dispname.csv`` mapping where present and "
        "fall back to IOC's *Family (English)* otherwise.\n"
    )


if __name__ == '__main__':
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument('--ioc', type=Path, required=True,
                   help='Path to IOC master XLSX (downloaded out-of-band).')
    p.add_argument('--alpha-codes-text', type=Path, required=True,
                   help='Path to pdftotext -layout output of IBP alpha codes PDF.')
    args = p.parse_args()
    build(args.ioc, args.alpha_codes_text)
